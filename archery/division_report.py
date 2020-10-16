import argparse
from dataclasses import dataclass
from datetime import datetime, timedelta
from operator import attrgetter
import os
import re
import subprocess
import sys

from openpyxl import load_workbook

FIRST_NAME_CELL = "First Name"
LAST_NAME_CELL = "Last Name"
DATE_OF_BIRTH_CELL = "DOB"
MENS_DIVISION_CELL = "Men's Divisions"
WOMENS_DIVISION_CELL = "Women's Divisions"

DIGITS_IN_DIVISION = re.compile(r"\d")


# Archer class
@dataclass
class Archer:
    first_name: str
    last_name: str
    divison_type: str
    divison_bow: str
    divison_class: str


# Reporting orders
DIVISION_TYPE_ORDER = ["Men's", "Women's"]
DIVISION_BOW_ORDER = ["Barebow", "Longbow", "Recurve", "Compound"]
DIVISION_CLASS_ORDER = [
    "Master",
    "Senior",
    "Junior",
    "Cadet",
    "Cub",
    "Bowman",
    "Yeoman",
]

AGE_DATE_FORMAT = "%Y-%m-%d"
DATE_INPUT_FORMAT = "%Y/%m/%d"
DATE_OUTPUT_FORMAT = "%B %d, %Y"
DATE_FILE_FORMAT = "%Y%m%d"


def masters_age_bracket(age):
    if age < 50:
        return None
    elif age >= 50 and age < 60:
        return 50
    elif age >= 60 and age < 70:
        return 60
    else:
        return 70


def main(opts):
    try:
        ifile = os.path.expanduser(os.path.expandvars(opts.report_file))
    except IndexError:
        print("Usage: division_report.py <Input Spreadsheet>")

    if opts.date is None:
        tournament_date = datetime.now()
    else:
        tournament_date = datetime.strptime(opts.date, DATE_INPUT_FORMAT)

    workbook = load_workbook(ifile)
    sheet = workbook.active

    # Find column indexes
    title_row = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    FIRST_NAME = title_row.index(FIRST_NAME_CELL)
    LAST_NAME = title_row.index(LAST_NAME_CELL)
    DATE_OF_BIRTH = title_row.index(DATE_OF_BIRTH_CELL)
    MENS_DIVISION = title_row.index(MENS_DIVISION_CELL)
    WOMENS_DIVISION = title_row.index(WOMENS_DIVISION_CELL)

    if opts.masters_age:
        DIVISION_CLASS_ORDER.pop(0)
        DIVISION_CLASS_ORDER.insert(0, "Master 50")
        DIVISION_CLASS_ORDER.insert(0, "Master 60")
        DIVISION_CLASS_ORDER.insert(0, "Master 70")

    archers = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if opts.verbose:
            print(f"Row: {row}")
        if row[FIRST_NAME] is None:
            continue
        if row[MENS_DIVISION] is not None:
            divison_type = "Men's"
            division = row[MENS_DIVISION]
        if row[WOMENS_DIVISION] is not None:
            divison_type = "Women's"
            division = row[WOMENS_DIVISION]

        if "(" in division:
            if DIGITS_IN_DIVISION.search(division) is not None:
                divison_bow, _, divison_class, _ = division.split()
            else:
                divison_bow, _, divison_class = division.split()
        elif DIGITS_IN_DIVISION.search(division) is not None:
            divison_bow, divison_class, _ = division.split()
        else:
            divison_bow, divison_class = division.split()

        if opts.masters_age:
            dob_date = datetime.strptime(row[DATE_OF_BIRTH], AGE_DATE_FORMAT)
            age_cutoff_date = datetime(tournament_date.year, 1, 1)
            age_at_tournament = age_cutoff_date.year - dob_date.year
            age_bracket = masters_age_bracket(age_at_tournament)
            if age_bracket is not None:
                divison_class = f"{divison_class} {age_bracket}"

        archers.append(
            Archer(
                first_name=row[FIRST_NAME].title(),
                last_name=row[LAST_NAME].title(),
                divison_type=divison_type,
                divison_bow=divison_bow,
                divison_class=divison_class,
            )
        )

    if opts.tournament is None:
        tournament = "Tournament"
    else:
        tournament = opts.tournament

    if opts.ndays > 1:
        tournament_end_date = tournament_date + timedelta(days=opts.ndays - 1)
    else:
        tournament_end_date = None

    report_lines = []
    report_lines.append("---" + os.linesep)
    report_lines.append(f"title: {tournament}" + os.linesep)
    report_lines.append("..." + os.linesep)
    if tournament_end_date is None:
        formatted_date = tournament_date.strftime(DATE_OUTPUT_FORMAT)
    else:
        if tournament_end_date.month == tournament_date.month:
            start_parts = tournament_date.strftime(DATE_OUTPUT_FORMAT).split()
            end_parts = tournament_end_date.strftime(DATE_OUTPUT_FORMAT).split()
            formatted_date = f"{start_parts[0]} {start_parts[1].strip(',')}-{end_parts[1]} {start_parts[2]}"
        if tournament_end_date.month > tournament_date.month:
            start_parts = tournament_date.strftime(DATE_OUTPUT_FORMAT).split()
            end_parts = tournament_end_date.strftime(DATE_OUTPUT_FORMAT).split()
            formatted_date = (
                f"{start_parts[0]} {start_parts[1].strip(',')}-"
                "{end_parts[0]} {end_parts[1]} {start_parts[2]}"
            )
    report_lines.append(f"{formatted_date}" + os.linesep)
    report_lines.append(os.linesep)

    report_lines.append(f"Number of Registered Archers: {len(archers)}" + os.linesep)
    report_lines.append(os.linesep)

    if opts.reverse:
        DIVISION_TYPE_ORDER.reverse()

    for dto in DIVISION_TYPE_ORDER:
        for dbo in DIVISION_BOW_ORDER:
            for dco in DIVISION_CLASS_ORDER:
                class_archers = []
                for archer in archers:
                    if (
                        archer.divison_type == dto
                        and archer.divison_bow == dbo
                        and archer.divison_class == dco
                    ):
                        class_archers.append(archer)
                if class_archers:
                    report_list = sorted(class_archers, key=attrgetter("last_name"))
                    report_lines.append(f"### {dto} {dbo} {dco}" + os.linesep)
                    for class_archer in report_list:
                        full_name = (
                            f"{class_archer.first_name} {class_archer.last_name}"
                        )
                        report_lines.append(f"{full_name}  " + os.linesep)
                    report_lines.append(os.linesep)

    values = [x.lower() for x in tournament.split(" ")]
    values.append(tournament_date.strftime(DATE_FILE_FORMAT))
    file_name = "_".join(values)

    markdown_file = f"{file_name}.md"
    with open(markdown_file, "w") as rfile:
        rfile.writelines(report_lines)

    pdf_file = markdown_file.split(".")[0] + ".pdf"

    subprocess.run(
        ["pandoc", "-s", "--pdf-engine", "weasyprint", "-o", pdf_file, markdown_file]
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "report_file",
        help="The Excel spreadsheet containing the USA Archery information",
    )
    parser.add_argument(
        "-t",
        "--tournament",
        dest="tournament",
        help="The name of the archery tournament",
    )
    parser.add_argument(
        "-d",
        "--date",
        dest="date",
        help="The date of the tournament in YYYY/MM/DD format.",
    )

    parser.add_argument(
        "-n",
        "--n-days",
        dest="ndays",
        type=int,
        default=1,
        help="Number of days for the tournament.",
    )

    parser.add_argument(
        "-r",
        "--reverse",
        dest="reverse",
        action="store_true",
        help="Reverse order of division type.",
    )

    parser.add_argument(
        "--masters-age",
        dest="masters_age",
        action="store_true",
        help="Sort Master's divisions based on age.",
    )

    parser.add_argument(
        "-v",
        "--verbose",
        dest="verbose",
        action="store_true",
        help="Print out intermediate information.",
    )

    args = parser.parse_args()

    main(args)
